import openpyxl
from openpyxl.utils.cell import coordinate_from_string, absolute_coordinate
from openpyxl.xml.constants import MAX_COLUMN, MAX_ROW, MIN_COLUMN


print("Opening Excel and reading file......")
# wb = openpyxl.load_workbook(r"C:\Python Stuff\Test Log.xlsx")
wb = openpyxl.load_workbook(r"D:\Python Stuff\Test Log.xlsx")
time_sheet = wb.worksheets[9]

ws = wb.worksheets[0]
start_cell = 10


names = wb.sheetnames
names.remove("Time")
names.remove("Set up Checklist")
names.remove("Extra")


# print("Current names to scan.... ")
# print(names)

print("Scanning cells.......")
for sheet in names:
    print(sheet) # Prints all sheet names to cycle through
    for row in ws.iter_rows(min_row=12, max_row=30, min_col=3, max_col=3, values_only=True): # Goes through each row in worksheet
        for model_number in row: # Gets model from cell
            model = str(model_number).lower()  # Makes model into string and lowercase to compare
            for row in time_sheet.iter_rows(max_col=1): # Goes to time sheet and scans rows
                for location in row: # Gets model from rom
                    model_time = str(location).lower() # Converts model to string and lowercase
                    print(location.absolute_coordinate)
                    if model == model_time: # Compares both models together
                        
                        print("yes")

                
