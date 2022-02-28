import openpyxl
from openpyxl.utils.cell import coordinate_from_string, absolute_coordinate
from openpyxl.xml.constants import MAX_COLUMN, MAX_ROW, MIN_COLUMN


print("Opening Excel and reading file......")
# wb = openpyxl.load_workbook(r"C:\Python Stuff\Test Log.xlsx")
wb = openpyxl.load_workbook(r"D:\Python Stuff\Test Log.xlsx")
time_sheet = wb.worksheets[9]

ws = wb.worksheets[0]
start_cell = 10


names = wb.sheetnames
names.remove("Time")
names.remove("Set up Checklist")
names.remove("Extra")
